from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional

from openpyxl import load_workbook


@dataclass(frozen=True)
class CategoryBook:
    expense_categories: Dict[str, List[str]]
    income_categories: List[str]

    @classmethod
    def from_workbook(cls, workbook_path: Path) -> "CategoryBook":
        workbook = load_workbook(workbook_path, data_only=False, read_only=True)
        if "Справочники" not in workbook.sheetnames:
            raise ValueError("Workbook must contain 'Справочники' sheet")

        sheet = workbook["Справочники"]
        expense_categories: Dict[str, List[str]] = {}
        income_categories: List[str] = []

        for column in range(1, sheet.max_column + 1):
            header = _cell_text(sheet.cell(1, column).value)
            if not header:
                continue

            values = [
                value
                for row in range(2, sheet.max_row + 1)
                if (value := _cell_text(sheet.cell(row, column).value))
            ]
            if header == "Доход":
                income_categories = values
            elif values:
                expense_categories[header] = values

        return cls(expense_categories=expense_categories, income_categories=income_categories)

    def is_valid_expense(self, category: Optional[str], subcategory: Optional[str]) -> bool:
        if not category or not subcategory:
            return False
        return subcategory in self.expense_categories.get(category, [])

    def is_valid_income(self, category: Optional[str]) -> bool:
        if not category:
            return False
        return category in self.income_categories

    def expense_prompt_text(self) -> str:
        rows = []
        for category, subcategories in self.expense_categories.items():
            if subcategories:
                rows.append(f"{category}: {', '.join(subcategories)}")
        return "\n".join(rows)

    def income_prompt_text(self) -> str:
        return ", ".join(self.income_categories)

    def has_subcategory(self, category: str, subcategory: str) -> bool:
        return subcategory in self.expense_categories.get(category, [])


def _cell_text(value: object) -> str:
    return str(value).strip() if value is not None else ""


def build_keyword_rules() -> Dict[str, tuple[str, str]]:
    return {
        "burger king": ("Еда", "Фастфуд"),
        "бургер кинг": ("Еда", "Фастфуд"),
        "kfc": ("Еда", "Фастфуд"),
        "вкусно и точка": ("Еда", "Фастфуд"),
        "вкусно — и точка": ("Еда", "Фастфуд"),
        "вкусно - и точка": ("Еда", "Фастфуд"),
        "макдональдс": ("Еда", "Фастфуд"),
        "skuratov": ("Еда", "Кофейни"),
        "surf coffee": ("Еда", "Кофейни"),
        "cofix": ("Еда", "Кофейни"),
        "кофикс": ("Еда", "Кофейни"),
        "метрополитен": ("Транспорт", "Местный транспорт"),
        "metro moscow": ("Транспорт", "Местный транспорт"),
        "ржд": ("Транспорт", "Местный транспорт"),
        "yandex fasten": ("Транспорт", "Такси"),
        "яндекс такси": ("Транспорт", "Такси"),
        "самокаты - яндекс": ("Транспорт", "Самокаты"),
        "самокаты — яндекс": ("Транспорт", "Самокаты"),
        "yandex go": ("Транспорт", "Самокаты"),
        "самокат": ("Еда", "Супермаркеты"),
        "пятерочка": ("Еда", "Супермаркеты"),
        "пятёрочка": ("Еда", "Супермаркеты"),
        "магнит": ("Еда", "Супермаркеты"),
        "чижик": ("Еда", "Супермаркеты"),
        "spar": ("Еда", "Супермаркеты"),
        "спар": ("Еда", "Супермаркеты"),
        "перекрёсток": ("Еда", "Супермаркеты"),
        "перекресток": ("Еда", "Супермаркеты"),
        "августина": ("Еда", "Супермаркеты"),
        "авито": ("Шоппинг", "Маркетплейсы"),
        "озон": ("Шоппинг", "Маркетплейсы"),
        "wildberries": ("Шоппинг", "Маркетплейсы"),
        "wb": ("Шоппинг", "Маркетплейсы"),
        "kvaligate": ("Досуг", "Цифровые товары"),
        "цифровые товары": ("Досуг", "Цифровые товары"),
        "good vibes": ("Остальное", "Разное"),
        "табак": ("Остальное", "Разное"),
        "аренда": ("Дом", "Аренда"),
        "квартира": ("Дом", "Аренда"),
        "съем": ("Дом", "Аренда"),
        "съём": ("Дом", "Аренда"),
        "домашние вещи": ("Дом", "Вещи"),
        "для дома": ("Дом", "Вещи"),
    }


def apply_keyword_rules(name: str, category_book: CategoryBook) -> Optional[tuple[str, str]]:
    normalized = name.casefold()
    for keyword, mapping in build_keyword_rules().items():
        category, subcategory = mapping
        if keyword in normalized and category_book.has_subcategory(category, subcategory):
            return mapping
    return None


def missing_required_subcategories(category_book: CategoryBook) -> List[str]:
    required = [
        ("Еда", "Фастфуд"),
        ("Еда", "Рестораны"),
        ("Еда", "Супермаркеты"),
        ("Транспорт", "Такси"),
        ("Транспорт", "Местный транспорт"),
    ]
    return [
        f"{category} / {subcategory}"
        for category, subcategory in required
        if not category_book.has_subcategory(category, subcategory)
    ]
