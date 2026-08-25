from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from .excel_writer import EXPENSE_SHEET, EXPENSE_TABLE, INCOME_SHEET, INCOME_TABLE
from .models import OperationType
from .storage import Storage


class ExcelReportReader:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path

    def read_budget_entries(self) -> List[Dict[str, Any]]:
        workbook = load_workbook(self.workbook_path, data_only=False)
        entries: List[Dict[str, Any]] = []
        entries.extend(_read_expense_entries(workbook[EXPENSE_SHEET], EXPENSE_TABLE))
        entries.extend(_read_income_entries(workbook[INCOME_SHEET], INCOME_TABLE))
        return entries

    def sync_to_storage(self, storage: Storage) -> int:
        return storage.replace_budget_entries(self.read_budget_entries())


def _read_expense_entries(sheet, table_name: str) -> Iterable[Dict[str, Any]]:
    for row in _table_data_rows(sheet, table_name):
        operation_date = _parse_cell_date(sheet.cell(row, 3).value)
        amount = _parse_amount(sheet.cell(row, 7).value)
        if operation_date is None or amount is None:
            continue
        yield {
            "source": "excel_sync",
            "operation_hash": None,
            "workbook_sheet": EXPENSE_SHEET,
            "workbook_row": row,
            "operation_date": operation_date.isoformat(),
            "operation_type": OperationType.EXPENSE.value,
            "amount": abs(amount),
            "category": _cell_text(sheet.cell(row, 4).value),
            "subcategory": _cell_text(sheet.cell(row, 5).value),
            "name": _cell_text(sheet.cell(row, 6).value),
            "note": _cell_text(sheet.cell(row, 8).value),
            "bank": "",
        }


def _read_income_entries(sheet, table_name: str) -> Iterable[Dict[str, Any]]:
    for row in _table_data_rows(sheet, table_name):
        operation_date = _parse_cell_date(sheet.cell(row, 3).value)
        amount = _parse_amount(sheet.cell(row, 7).value)
        if operation_date is None or amount is None:
            continue
        yield {
            "source": "excel_sync",
            "operation_hash": None,
            "workbook_sheet": INCOME_SHEET,
            "workbook_row": row,
            "operation_date": operation_date.isoformat(),
            "operation_type": OperationType.INCOME.value,
            "amount": abs(amount),
            "category": _cell_text(sheet.cell(row, 5).value),
            "subcategory": None,
            "name": _cell_text(sheet.cell(row, 4).value),
            "note": _cell_text(sheet.cell(row, 10).value),
            "bank": "",
        }


def _table_data_rows(sheet, table_name: str) -> range:
    table = sheet.tables[table_name]
    _min_col, min_row, _max_col, max_row = range_boundaries(table.ref)
    return range(min_row + 1, max_row + 1)


def _parse_cell_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip() if value is not None else ""
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        return None


def _parse_amount(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(" ", "").replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _cell_text(value: Any) -> str:
    return str(value).strip() if value is not None else ""
