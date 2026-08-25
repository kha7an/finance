from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from .excel_writer import EXPENSE_SHEET, EXPENSE_TABLE, INCOME_SHEET, INCOME_TABLE
from .models import OperationType
from .storage import Storage


class ExcelExporter:
    def __init__(self, storage: Storage, export_dir: Path) -> None:
        self.storage = storage
        self.export_dir = export_dir

    def export(self, owner_id: str, start_date: date, end_date: date) -> Path:
        with self.storage.owner_scope(owner_id):
            entries = self.storage.all_budget_entries(start_date, end_date)
            category_book = self.storage.category_book()

        workbook = Workbook()
        default = workbook.active
        workbook.remove(default)

        expenses = workbook.create_sheet(EXPENSE_SHEET)
        incomes = workbook.create_sheet(INCOME_SHEET)
        refs = workbook.create_sheet("Справочники")

        _write_expenses(expenses, [entry for entry in entries if entry["operation_type"] == OperationType.EXPENSE.value])
        _write_incomes(incomes, [entry for entry in entries if entry["operation_type"] == OperationType.INCOME.value])
        _write_refs(refs, category_book.expense_categories, category_book.income_categories)

        owner_dir = self.export_dir / _safe_owner_id(owner_id)
        owner_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        path = owner_dir / f"budget-{timestamp}.xlsx"
        workbook.save(path)
        return path


def _write_expenses(sheet, entries: List[Dict[str, Any]]) -> None:
    sheet.append(["Месяц", "Год", "Дата", "Категория", "Подкатегория", "Подробное описание", "Стоимость", "Комментарий"])
    rows = entries or [_blank_expense_entry()]
    for index, entry in enumerate(rows, start=2):
        sheet.append(
            [
                f'=TEXT(C{index},"ММММ")',
                f"=YEAR(C{index})",
                _entry_date(entry),
                entry.get("category") or "",
                entry.get("subcategory") or "",
                entry.get("name") or "",
                float(entry.get("amount") or 0),
                entry.get("note") or _comment(entry),
            ]
        )
    _add_table(sheet, EXPENSE_TABLE, f"A1:H{max(2, len(rows) + 1)}")


def _write_incomes(sheet, entries: List[Dict[str, Any]]) -> None:
    sheet.append(["Месяц", "Год", "Дата", "От кого пришло", "Категория", "Подробное описание", "Приход", "Вложено в бюджет", "Остаток на накопления", "Комментарий"])
    rows = entries or [_blank_income_entry()]
    for index, entry in enumerate(rows, start=2):
        sheet.append(
            [
                f'=TEXT(C{index},"ММММ")',
                f"=YEAR(C{index})",
                _entry_date(entry),
                entry.get("name") or "",
                entry.get("category") or "",
                entry.get("note") or "",
                float(entry.get("amount") or 0),
                None,
                f"=УчетДоходов[[#This Row],[Приход]]-УчетДоходов[[#This Row],[Вложено в бюджет]]",
                _comment(entry),
            ]
        )
    _add_table(sheet, INCOME_TABLE, f"A1:J{max(2, len(rows) + 1)}")


def _write_refs(sheet, expense_categories: Dict[str, List[str]], income_categories: List[str]) -> None:
    headers = list(expense_categories.keys()) + ["Доход"]
    sheet.append(headers)
    max_rows = max([len(items) for items in expense_categories.values()] + [len(income_categories), 1])
    for index in range(max_rows):
        row = []
        for subcategories in expense_categories.values():
            row.append(subcategories[index] if index < len(subcategories) else "")
        row.append(income_categories[index] if index < len(income_categories) else "")
        sheet.append(row)


def _add_table(sheet, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    sheet.add_table(table)


def _entry_date(entry: Dict[str, Any]) -> Optional[date]:
    value = entry.get("operation_date")
    if isinstance(value, date):
        return value
    if not value:
        return None
    return date.fromisoformat(str(value))


def _comment(entry: Dict[str, Any]) -> str:
    bank = str(entry.get("bank") or "").strip()
    name = str(entry.get("name") or "").strip()
    if bank and name:
        return f"auto: {bank} / {name}"
    return f"auto: {bank}" if bank else ""


def _blank_expense_entry() -> Dict[str, Any]:
    return {"operation_date": None, "amount": 0}


def _blank_income_entry() -> Dict[str, Any]:
    return {"operation_date": None, "amount": 0}


def _safe_owner_id(owner_id: str) -> str:
    return "".join(ch if ch.isalnum() or ch in {"-", "_"} else "_" for ch in owner_id)
