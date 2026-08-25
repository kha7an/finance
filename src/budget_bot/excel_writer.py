from __future__ import annotations

import copy
import threading
from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries
from openpyxl.worksheet.table import Table

from .models import OperationType, ParsedOperation


EXPENSE_SHEET = "Учет расходов"
EXPENSE_TABLE = "УТ_Данные"
INCOME_SHEET = "Учет доходов"
INCOME_TABLE = "УчетДоходов"


@dataclass(frozen=True)
class WriteResult:
    sheet_name: str
    row: int


@dataclass(frozen=True)
class DeleteResult:
    shifted_rows: bool


class ExcelWriter:
    def __init__(self, workbook_path: Path) -> None:
        self.workbook_path = workbook_path
        self._lock = threading.Lock()

    def append_operation(self, operation: ParsedOperation, bank: str) -> WriteResult:
        with self._lock:
            if operation.type == OperationType.EXPENSE:
                return self._append_expense(operation, bank)
            if operation.type == OperationType.INCOME:
                return self._append_income(operation, bank)
            raise ValueError(f"Cannot write operation type {operation.type.value!r} to Excel")

    def reset_workbook(self) -> None:
        with self._lock:
            workbook = load_workbook(self.workbook_path, data_only=False)
            _reset_table_sheet(workbook[EXPENSE_SHEET], workbook[EXPENSE_SHEET].tables[EXPENSE_TABLE], max_col=8)
            _reset_table_sheet(workbook[INCOME_SHEET], workbook[INCOME_SHEET].tables[INCOME_TABLE], max_col=10)
            workbook.save(self.workbook_path)

    def update_entry(self, sheet_name: str, row: int, operation: ParsedOperation, bank: str) -> None:
        with self._lock:
            workbook = load_workbook(self.workbook_path, data_only=False)
            if sheet_name == EXPENSE_SHEET:
                _write_expense_row(workbook[sheet_name], row, operation, bank)
            elif sheet_name == INCOME_SHEET:
                _write_income_row(workbook[sheet_name], row, operation, bank)
            else:
                raise ValueError(f"Unknown budget sheet: {sheet_name!r}")
            workbook.save(self.workbook_path)

    def delete_entry(self, sheet_name: str, row: int) -> DeleteResult:
        with self._lock:
            workbook = load_workbook(self.workbook_path, data_only=False)
            if sheet_name == EXPENSE_SHEET:
                sheet = workbook[EXPENSE_SHEET]
                table = sheet.tables[EXPENSE_TABLE]
                max_col = 8
            elif sheet_name == INCOME_SHEET:
                sheet = workbook[INCOME_SHEET]
                table = sheet.tables[INCOME_TABLE]
                max_col = 10
            else:
                raise ValueError(f"Unknown budget sheet: {sheet_name!r}")

            _min_col, min_row, _max_col, max_row = range_boundaries(table.ref)
            if row <= min_row or row > max_row:
                raise ValueError(f"Row {row} is outside table {table.ref}")
            if max_row <= min_row + 1:
                _clear_row(sheet, row, max_col)
                shifted_rows = False
            else:
                sheet.delete_rows(row, 1)
                _set_table_end_row(table, max(max_row - 1, min_row + 1))
                shifted_rows = True
            workbook.save(self.workbook_path)
            return DeleteResult(shifted_rows=shifted_rows)

    def _append_expense(self, operation: ParsedOperation, bank: str) -> WriteResult:
        workbook = load_workbook(self.workbook_path, data_only=False)
        sheet = workbook[EXPENSE_SHEET]
        table = sheet.tables[EXPENSE_TABLE]
        row = _next_table_row(table)
        previous_row = row - 1

        _copy_row_style(sheet, previous_row, row, 8)
        _write_expense_row(sheet, row, operation, bank)

        _extend_table(table, row)
        workbook.save(self.workbook_path)
        return WriteResult(sheet_name=EXPENSE_SHEET, row=row)

    def _append_income(self, operation: ParsedOperation, bank: str) -> WriteResult:
        workbook = load_workbook(self.workbook_path, data_only=False)
        sheet = workbook[INCOME_SHEET]
        table = sheet.tables[INCOME_TABLE]
        row = _next_table_row(table)
        previous_row = row - 1

        _copy_row_style(sheet, previous_row, row, 10)
        _write_income_row(sheet, row, operation, bank)

        _extend_table(table, row)
        workbook.save(self.workbook_path)
        return WriteResult(sheet_name=INCOME_SHEET, row=row)


def _next_table_row(table: Table) -> int:
    _, end = table.ref.split(":")
    return int("".join(ch for ch in end if ch.isdigit())) + 1


def _extend_table(table: Table, last_row: int) -> None:
    start, end = table.ref.split(":")
    end_col = "".join(ch for ch in end if ch.isalpha())
    table.ref = f"{start}:{end_col}{last_row}"


def _set_table_end_row(table: Table, last_row: int) -> None:
    min_col, min_row, max_col, _max_row = range_boundaries(table.ref)
    table.ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{last_row}"
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref


def _copy_row_style(sheet, source_row: int, target_row: int, max_col: int) -> None:
    for column in range(1, max_col + 1):
        source = sheet.cell(source_row, column)
        target = sheet.cell(target_row, column)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.font:
            target.font = copy.copy(source.font)
        if source.fill:
            target.fill = copy.copy(source.fill)
        if source.border:
            target.border = copy.copy(source.border)


def _reset_table_sheet(sheet, table: Table, max_col: int) -> None:
    min_col, _min_row, _max_col, _max_row = range_boundaries(table.ref)
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)
    for column in range(min_col, min_col + max_col):
        sheet.cell(2, column).value = None
    table.ref = f"{sheet.cell(1, min_col).coordinate}:{sheet.cell(2, min_col + max_col - 1).coordinate}"
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref
    table.sortState = None


def _clear_row(sheet, row: int, max_col: int) -> None:
    for column in range(1, max_col + 1):
        sheet.cell(row, column).value = None


def _write_expense_row(sheet, row: int, operation: ParsedOperation, bank: str) -> None:
    sheet.cell(row, 1).value = f'=TEXT(C{row},"ММММ")'
    sheet.cell(row, 2).value = f"=YEAR(C{row})"
    sheet.cell(row, 3).value = operation.date
    sheet.cell(row, 4).value = operation.category
    sheet.cell(row, 5).value = operation.subcategory
    sheet.cell(row, 6).value = operation.note or operation.name
    sheet.cell(row, 7).value = operation.excel_amount
    sheet.cell(row, 8).value = _comment(bank, operation)


def _write_income_row(sheet, row: int, operation: ParsedOperation, bank: str) -> None:
    sheet.cell(row, 1).value = f'=TEXT(C{row},"ММММ")'
    sheet.cell(row, 2).value = f"=YEAR(C{row})"
    sheet.cell(row, 3).value = operation.date
    sheet.cell(row, 4).value = operation.name
    sheet.cell(row, 5).value = operation.category
    sheet.cell(row, 6).value = operation.note
    sheet.cell(row, 7).value = operation.excel_amount
    sheet.cell(row, 8).value = None
    sheet.cell(row, 9).value = (
        f"=УчетДоходов[[#This Row],[Приход]]"
        f"-УчетДоходов[[#This Row],[Вложено в бюджет]]"
    )
    sheet.cell(row, 10).value = _comment(bank, operation)


def _comment(bank: str, operation: ParsedOperation) -> str:
    name = operation.name.strip()
    if name:
        return f"auto: {bank} / {name}"
    return f"auto: {bank}"
