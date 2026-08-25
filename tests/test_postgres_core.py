from __future__ import annotations

import os
import subprocess
import sys
from datetime import date
from pathlib import Path
from types import SimpleNamespace
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from budget_bot.app_factory import AppContext
from budget_bot.config import Settings
from budget_bot.excel_exporter import ExcelExporter
from budget_bot.models import OperationStatus, OperationType, ParsedOperation, ParsedScreenshot
from budget_bot.processor import OperationDecision, ProcessingResult, ScreenshotProcessor
from budget_bot.storage import Storage, operation_hash
from budget_bot.telegram_bot import (
    TELEGRAM_POLLING_CONFLICT_SLEEP_SECONDS,
    TELEGRAM_POLLING_ERROR_SLEEP_SECONDS,
    TelegramApiError,
    TelegramBot,
)


DATABASE_URL = os.getenv("DATABASE_URL", "postgresql://budget_bot:budget_bot@127.0.0.1:5433/budget_bot")


def owner_id() -> str:
    return f"test:{uuid4().hex}"


def make_operation(
    operation_type: OperationType = OperationType.EXPENSE,
    amount: float = -709.96,
    category: str | None = "Еда",
    subcategory: str | None = "Фастфуд",
    name: str = "Бургер Кинг",
) -> ParsedOperation:
    return ParsedOperation(
        date=date(2026, 8, 21),
        name=name,
        amount=amount,
        type=operation_type,
        category=category,
        subcategory=subcategory,
    )


def test_postgres_bootstrap_creates_core_tables_and_indexes() -> None:
    storage = Storage(DATABASE_URL)
    expected_tables = {
        "users",
        "categories",
        "subcategories",
        "source_images",
        "operations",
        "budget_entries",
        "pending_actions",
        "learned_expense_categories",
        "telegram_chats",
        "reminder_settings",
        "reminder_deliveries",
    }
    expected_indexes = {
        "source_images_owner_hash_idx",
        "operations_owner_hash_idx",
        "budget_entries_owner_date_type_idx",
    }

    with storage._connect() as connection:
        tables = {
            row["table_name"]
            for row in connection.execute(
                """
                SELECT table_name
                FROM information_schema.tables
                WHERE table_schema = 'public'
                """
            ).fetchall()
        }
        indexes = {
            row["indexname"]
            for row in connection.execute(
                """
                SELECT indexname
                FROM pg_indexes
                WHERE schemaname = 'public'
                """
            ).fetchall()
        }

    assert expected_tables <= tables
    assert expected_indexes <= indexes


def test_postgres_bootstrap_creates_default_categories() -> None:
    storage = Storage(DATABASE_URL)
    owner = owner_id()

    with storage.owner_scope(owner):
        book = storage.category_book()

    assert book.is_valid_expense("Еда", "Фастфуд")
    assert book.is_valid_expense("Транспорт", "Такси")
    assert book.is_valid_income("Зарплата")


def test_processor_writes_auto_expense_to_postgres_and_isolates_owners() -> None:
    storage = Storage(DATABASE_URL)
    first_owner = owner_id()
    second_owner = owner_id()
    parsed = ParsedScreenshot.from_json(
        {
            "bank": "tbank",
            "operations": [
                {
                    "date": "2026-08-21",
                    "name": "Бургер Кинг",
                    "amount": -709.96,
                    "type": "expense",
                    "category": "Еда",
                    "subcategory": "Фастфуд",
                }
            ],
        }
    )

    with storage.owner_scope(first_owner):
        processor = ScreenshotProcessor(storage=storage, category_book=storage.category_book())
        first = processor.process(b"same-image", parsed)
        duplicate = processor.process(b"same-image", parsed)
        first_summary = storage.expense_summary(date(2026, 8, 1), date(2026, 8, 31))

    with storage.owner_scope(second_owner):
        processor = ScreenshotProcessor(storage=storage, category_book=storage.category_book())
        second = processor.process(b"same-image", parsed)
        second_summary = storage.expense_summary(date(2026, 8, 1), date(2026, 8, 31))

    assert [item.status for item in first.decisions] == [OperationStatus.AUTO_WRITTEN]
    assert duplicate.decisions == []
    assert [item.status for item in second.decisions] == [OperationStatus.AUTO_WRITTEN]
    assert first_summary["count"] == 1
    assert second_summary["count"] == 1


def test_learned_category_rules_are_owner_scoped() -> None:
    storage = Storage(DATABASE_URL)
    first_owner = owner_id()
    second_owner = owner_id()

    with storage.owner_scope(first_owner):
        storage.save_expense_category_rule("Mystery Place", "Еда", "Фастфуд")
        first_rule = storage.get_expense_category_rule("Mystery Place")

    with storage.owner_scope(second_owner):
        second_rule = storage.get_expense_category_rule("Mystery Place")

    assert first_rule == ("Еда", "Фастфуд")
    assert second_rule is None


def test_pending_actions_and_operation_rows_are_owner_scoped() -> None:
    storage = Storage(DATABASE_URL)
    first_owner = owner_id()
    second_owner = owner_id()
    operation = make_operation(category=None, subcategory=None)

    with storage.owner_scope(first_owner):
        op_hash = operation_hash("tbank", operation)
        storage.record_operation(op_hash, "image-1", "tbank", operation, OperationStatus.PENDING)
        storage.add_pending_action(op_hash, chat_id=100, message_id=200, prompt="Выбери категорию")
        first_pending = storage.get_pending_action(op_hash, 100)
        first_row = storage.get_operation(op_hash)

    with storage.owner_scope(second_owner):
        same_hash = operation_hash("tbank", operation)
        storage.record_operation(same_hash, "image-1", "tbank", operation, OperationStatus.PENDING)
        second_pending_before = storage.get_pending_action(same_hash, 100)
        second_row = storage.get_operation(same_hash)

    assert first_pending is not None
    assert first_row is not None
    assert second_pending_before is None
    assert second_row is not None
    assert first_row["operation_hash"] != second_row["operation_hash"]


def test_budget_entry_stats_edit_delete_and_reset_are_owner_scoped() -> None:
    storage = Storage(DATABASE_URL)
    first_owner = owner_id()
    second_owner = owner_id()
    operation = make_operation()

    with storage.owner_scope(first_owner):
        entry_id = storage.append_budget_entry("bot", operation_hash("tbank", operation), operation, "tbank")
        storage.update_budget_entry(entry_id, make_operation(amount=-1000, name="Updated"))
        edited = storage.get_budget_entry(entry_id)
        first_summary = storage.expense_summary(date(2026, 8, 1), date(2026, 8, 31))

    with storage.owner_scope(second_owner):
        other_id = storage.append_budget_entry("bot", operation_hash("tbank", operation), operation, "tbank")
        storage.delete_budget_entry(entry_id)
        other_still_exists = storage.get_budget_entry(other_id)
        storage.reset_all()
        other_after_reset = storage.get_budget_entry(other_id)

    with storage.owner_scope(first_owner):
        first_after_other_delete = storage.get_budget_entry(entry_id)
        first_after_other_reset = storage.expense_summary(date(2026, 8, 1), date(2026, 8, 31))
        storage.delete_budget_entry(entry_id)
        first_after_delete = storage.expense_summary(date(2026, 8, 1), date(2026, 8, 31))

    assert edited is not None
    assert edited["name"] == "Updated"
    assert edited["amount"] == 1000
    assert first_summary["total"] == 1000
    assert other_still_exists is not None
    assert other_after_reset is None
    assert first_after_other_delete is not None
    assert first_after_other_reset["count"] == 1
    assert first_after_delete["count"] == 0


def test_processor_pending_manual_expense_decision_writes_budget_entry() -> None:
    storage = Storage(DATABASE_URL)
    owner = owner_id()
    parsed = ParsedScreenshot.from_json(
        {
            "bank": "tbank",
            "operations": [
                {
                    "date": "2026-08-21",
                    "name": "Непонятная покупка",
                    "amount": -500,
                    "type": "expense",
                    "category": "Unknown",
                    "subcategory": "Unknown",
                }
            ],
        }
    )

    with storage.owner_scope(owner):
        processor = ScreenshotProcessor(storage=storage, category_book=storage.category_book())
        result = processor.process(b"pending-image", parsed)
        op_hash = operation_hash("tbank", result.decisions[0].operation)
        row = storage.get_operation(op_hash)
        bot = FakeTelegramBot(make_context(storage, owner, Path("data/exports")))
        bot._write_counted_operation(100, row, op_hash, "Еда", "Фастфуд")
        entries = storage.budget_entries(date(2026, 8, 1), date(2026, 8, 31))
        updated_row = storage.get_operation(op_hash)
        learned_rule = storage.get_expense_category_rule("Непонятная покупка")

    assert [decision.status for decision in result.decisions] == [OperationStatus.PENDING]
    assert len(entries) == 1
    assert entries[0]["category"] == "Еда"
    assert entries[0]["subcategory"] == "Фастфуд"
    assert updated_row["status"] == OperationStatus.AUTO_WRITTEN.value
    assert learned_rule == ("Еда", "Фастфуд")


def test_manual_income_and_expense_entries_write_to_postgres(tmp_path: Path) -> None:
    storage = Storage(DATABASE_URL)
    owner = owner_id()

    with storage.owner_scope(owner):
        bot = FakeTelegramBot(make_context(storage, owner, tmp_path))
        bot._manual_entry().write_operation(100, make_operation())
        bot._manual_entry().write_operation(
            100,
            make_operation(
                operation_type=OperationType.INCOME,
                amount=1250,
                category="Возврат",
                subcategory=None,
                name="Озон",
            ),
        )
        entries = storage.all_budget_entries(date(2026, 8, 1), date(2026, 8, 31))

    assert [entry["operation_type"] for entry in entries] == ["expense", "income"]
    assert [entry["source"] for entry in entries] == ["manual", "manual"]
    assert entries[0]["amount"] == 709.96
    assert entries[1]["amount"] == 1250


def test_excel_exporter_creates_current_workbook_format(tmp_path: Path) -> None:
    storage = Storage(DATABASE_URL)
    owner = owner_id()
    expense = ParsedOperation(
        date=date(2026, 8, 21),
        name="Бургер Кинг",
        amount=-709.96,
        type=OperationType.EXPENSE,
        category="Еда",
        subcategory="Фастфуд",
    )
    income = ParsedOperation(
        date=date(2026, 8, 22),
        name="Озон",
        amount=1250,
        type=OperationType.INCOME,
        category="Возврат",
    )

    with storage.owner_scope(owner):
        storage.append_budget_entry("bot", operation_hash("tbank", expense), expense, "tbank")
        storage.append_budget_entry("manual", operation_hash("manual", income), income, "manual")

    path = ExcelExporter(storage, tmp_path).export(owner, date(2026, 8, 1), date(2026, 8, 31))
    workbook = load_workbook(path, data_only=False)

    assert set(workbook.sheetnames) == {"Учет расходов", "Учет доходов", "Справочники"}
    expense_sheet = workbook["Учет расходов"]
    income_sheet = workbook["Учет доходов"]
    assert expense_sheet.tables["УТ_Данные"].ref == "A1:H2"
    assert income_sheet.tables["УчетДоходов"].ref == "A1:J2"
    assert expense_sheet["D2"].value == "Еда"
    assert expense_sheet["E2"].value == "Фастфуд"
    assert expense_sheet["G2"].value == 709.96
    assert income_sheet["E2"].value == "Возврат"
    assert income_sheet["I2"].value == "=УчетДоходов[[#This Row],[Приход]]-УчетДоходов[[#This Row],[Вложено в бюджет]]"


def test_telegram_export_sends_workbook_and_sync_is_deprecated(tmp_path: Path) -> None:
    storage = Storage(DATABASE_URL)
    owner = owner_id()

    with storage.owner_scope(owner):
        storage.append_budget_entry("bot", operation_hash("tbank", make_operation()), make_operation(), "tbank")
        bot = FakeTelegramBot(make_context(storage, owner, tmp_path))
        bot._handle_export_text(100, "/export 01.08-31.08")
        bot._handle_text({"chat": {"id": 100}, "from": {"id": 1}, "text": "/sync"})

    assert len(bot.sent_documents) == 1
    assert bot.sent_documents[0].exists()
    assert any("Синхронизации с Excel больше нет" in message["text"] for message in bot.sent_messages)


def test_telegram_polling_conflict_uses_longer_retry_delay() -> None:
    bot = TelegramBot.__new__(TelegramBot)

    conflict_delay = bot._polling_error_sleep_seconds(TelegramApiError("conflict", status_code=409))
    generic_delay = bot._polling_error_sleep_seconds(RuntimeError("network down"))

    assert conflict_delay == TELEGRAM_POLLING_CONFLICT_SLEEP_SECONDS
    assert generic_delay == TELEGRAM_POLLING_ERROR_SLEEP_SECONDS


def test_telegram_processing_result_shows_written_summary_with_pending_items(tmp_path: Path) -> None:
    written = make_operation(name="Пятерочка")
    pending = ParsedOperation(
        date=date(2026, 8, 1),
        name="Surf Coffee",
        amount=-340,
        type=OperationType.EXPENSE,
        category="Еда",
        subcategory="Кофе",
        occurrence_count=2,
        occurrence_confirmed=False,
    )
    result = ProcessingResult(
        image_hash="image-with-written-and-pending",
        bank="tbank",
        decisions=[
            OperationDecision(written, OperationStatus.AUTO_WRITTEN, "auto written", workbook_row=1),
            OperationDecision(pending, OperationStatus.PENDING, "same operation appears 2 times"),
        ],
    )
    pending_actions: list[dict[str, Any]] = []

    class FakeStorage:
        def add_pending_action(
            self,
            operation_hash: str,
            chat_id: int,
            message_id: int | None,
            prompt: str,
        ) -> None:
            pending_actions.append(
                {
                    "operation_hash": operation_hash,
                    "chat_id": chat_id,
                    "message_id": message_id,
                    "prompt": prompt,
                }
            )

    context = SimpleNamespace(
        settings=Settings(
            telegram_bot_token="test-token",
            telegram_allowed_user_ids={1},
            telegram_allow_all=False,
            telegram_api_base_url="https://api.telegram.org",
            telegram_file_base_url="https://api.telegram.org/file",
            telegram_proxy_url="",
            telegram_timeout_seconds=1,
            use_env_proxy=False,
            llm_provider="mock",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            openai_proxy_url="",
            openai_timeout_seconds=1,
            gemini_api_key="",
            gemini_model="gemini-2.5-flash",
            database_url=DATABASE_URL,
            google_credentials_path=Path("data/google-service-account.json"),
            export_dir=tmp_path,
            default_timezone="Europe/Moscow",
            reminder_enabled=False,
            reminder_default_time="21:00",
            api_token="",
            max_upload_bytes=10 * 1024 * 1024,
        ),
        storage=FakeStorage(),
    )
    bot = FakeTelegramBot(context)
    bot._send_processing_result(100, result)

    assert any(message["text"].startswith("Засчитано:") for message in bot.sent_messages)
    assert any("Решить: Surf Coffee" in message["text"] for message in bot.sent_messages)
    assert pending_actions


def test_cli_check_sync_mock_run_and_export_excel(tmp_path: Path) -> None:
    owner = owner_id()
    env = os.environ.copy()
    env.update(
        {
            "DATABASE_URL": DATABASE_URL,
            "PYTHONPATH": "src",
            "LLM_PROVIDER": "mock",
            "BUDGET_EXPORT_DIR": str(tmp_path),
        }
    )

    check = run_cli(env, "check")
    sync = run_cli(env, "sync")
    mock = run_cli(env, "mock-run", "--tag", f"pytest-{uuid4().hex}")
    export = run_cli(env, "export-excel", "--owner", owner, "--period", "01.08-31.08")

    assert "Category check: OK" in check.stdout
    assert "sync is deprecated; use export-excel" in sync.stdout
    assert "Bank: tbank" in mock.stdout
    assert "Excel export:" in export.stdout
    assert Path(export.stdout.strip().split("Excel export: ", 1)[1]).exists()


class FakeTelegramBot(TelegramBot):
    def __init__(self, context: AppContext) -> None:
        super().__init__(context)
        self.sent_messages: list[dict[str, Any]] = []
        self.sent_documents: list[Path] = []
        self.deleted_messages: list[tuple[int, int]] = []
        self.callback_answers: list[tuple[str, str]] = []

    def _send_message(
        self,
        chat_id: int,
        text: str,
        reply_markup: dict[str, Any] | None = None,
    ) -> int:
        message_id = len(self.sent_messages) + 1
        self.sent_messages.append(
            {
                "chat_id": chat_id,
                "text": text,
                "reply_markup": reply_markup,
                "message_id": message_id,
            }
        )
        return message_id

    def _send_document(self, chat_id: int, path: Path) -> bool:
        self.sent_documents.append(path)
        return True

    def _delete_message(self, chat_id: int, message_id: int) -> None:
        self.deleted_messages.append((chat_id, message_id))

    def _answer_callback(self, callback_id: str, text: str) -> None:
        self.callback_answers.append((callback_id, text))


def make_context(storage: Storage, owner: str, export_dir: Path) -> AppContext:
    return SimpleNamespace(
        settings=Settings(
            telegram_bot_token="test-token",
            telegram_allowed_user_ids={1},
            telegram_allow_all=False,
            telegram_api_base_url="https://api.telegram.org",
            telegram_file_base_url="https://api.telegram.org/file",
            telegram_proxy_url="",
            telegram_timeout_seconds=1,
            use_env_proxy=False,
            llm_provider="mock",
            openai_api_key="",
            openai_model="gpt-4o-mini",
            openai_proxy_url="",
            openai_timeout_seconds=1,
            gemini_api_key="",
            gemini_model="gemini-2.5-flash",
            database_url=DATABASE_URL,
            google_credentials_path=Path("data/google-service-account.json"),
            export_dir=export_dir,
            default_timezone="Europe/Moscow",
            reminder_enabled=False,
            reminder_default_time="21:00",
            api_token="",
            max_upload_bytes=10 * 1024 * 1024,
        ),
        storage=storage,
        owner_id=owner,
        category_book=storage.category_book(),
        budget_account_summary=lambda: f"Postgres: {DATABASE_URL.split('@')[-1]}",
    )


def run_cli(env: dict[str, str], *args: str) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        [sys.executable, "-m", "budget_bot.cli", *args],
        cwd=Path(__file__).resolve().parents[1],
        env=env,
        text=True,
        capture_output=True,
        check=True,
    )
